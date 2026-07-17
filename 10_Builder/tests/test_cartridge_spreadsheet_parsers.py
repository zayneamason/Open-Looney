from __future__ import annotations

import asyncio
import json
import sqlite3

import pytest

from luna.cartridge.builder import CartridgeBuilder
from luna.cartridge.parsers.csv import CSVParser


def _cells(nodes: list[dict]) -> list[dict]:
    return [node for node in nodes if node["type"] == "cell"]


def _meta(node: dict) -> dict:
    return node.get("meta") or {}


def test_csv_parser_detects_headers_and_preserves_blank_cells(tmp_path):
    source = tmp_path / "tasks.csv"
    source.write_text("Task,Owner,Status\nLaunch,Zayne,\nReview,,Done\n", encoding="utf-8")

    nodes = CSVParser().parse(source)
    rows = [node for node in nodes if node["type"] == "row"]
    cells = _cells(nodes)

    assert [node["type"] for node in nodes[:3]] == ["document", "section", "table"]
    assert _meta(rows[0])["header"] is True
    assert _meta(rows[1])["row_summary"] == "Task: Launch | Owner: Zayne"
    assert cells[5]["content"] == ""
    assert _meta(cells[5])["column_name"] == "Status"
    assert _meta(cells[5])["value_type"] == "text"


def test_csv_parser_handles_quoted_commas_and_newlines(tmp_path):
    source = tmp_path / "quoted.csv"
    source.write_text(
        'Name,Notes\n"Alice, A.","Line one\nLine two"\n',
        encoding="utf-8",
    )

    nodes = CSVParser().parse(source)
    cells = _cells(nodes)

    assert cells[2]["content"] == "Alice, A."
    assert cells[3]["content"] == "Line one\nLine two"
    assert _meta(cells[3])["address"] == "B2"


def test_csv_parser_generates_column_names_for_invalid_headers(tmp_path):
    source = tmp_path / "invalid_headers.csv"
    source.write_text("1,1,\nalpha,beta,gamma\n", encoding="utf-8")

    nodes = CSVParser().parse(source)
    table = next(node for node in nodes if node["type"] == "table")
    cells = _cells(nodes)

    assert _meta(table)["headers_from_source"] is False
    assert _meta(cells[3])["column_name"] == "Column A"
    assert _meta(cells[5])["column_name"] == "Column C"


def _add_defined_name(workbook, defined_name):
    if hasattr(workbook.defined_names, "add"):
        workbook.defined_names.add(defined_name)
    else:
        workbook.defined_names.append(defined_name)


def _write_workbook(path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    workbook = Workbook()
    tasks = workbook.active
    tasks.title = "Tasks"
    tasks.append(["Task", "Owner", "Status", "Score", "Total"])
    tasks.append(["Launch", "Zayne", "Backlog", 2, "=SUM(D2:D3)"])
    tasks.append(["Review", "Ada", "Done", 3, None])

    validation = DataValidation(
        type="list",
        formula1='"Backlog,In Progress,Done"',
        allow_blank=False,
    )
    tasks.add_data_validation(validation)
    validation.add("C2:C3")

    tasks.merge_cells("F1:G1")
    tasks["F1"] = "Merged Header"
    _add_defined_name(workbook, DefinedName("TaskNames", attr_text="'Tasks'!$A$2:$A$3"))

    lookup = workbook.create_sheet("Lookup")
    lookup.append(["Key", "Value"])
    lookup.append(["A", "Alpha"])
    workbook.save(path)


def test_spreadsheet_parser_preserves_workbook_logic(tmp_path):
    pytest.importorskip("openpyxl")
    from luna.cartridge.parsers.spreadsheet import SpreadsheetParser

    source = tmp_path / "tasks.xlsx"
    _write_workbook(source)

    nodes = SpreadsheetParser().parse(source)
    sections = [node for node in nodes if node["type"] == "section"]
    formula_cell = next(
        node for node in _cells(nodes)
        if _meta(node).get("formula") == "=SUM(D2:D3)"
    )
    status_cell = next(
        node for node in _cells(nodes)
        if _meta(node).get("address") == "Tasks!C2"
    )
    named_cell = next(
        node for node in _cells(nodes)
        if _meta(node).get("address") == "Tasks!A2"
    )
    merged_cell = next(
        node for node in _cells(nodes)
        if _meta(node).get("address") == "Tasks!F1"
    )

    assert [node["content"] for node in sections] == ["Tasks", "Lookup"]
    assert _meta(formula_cell)["formula_refs"] == ["Tasks!D2:D3"]
    assert _meta(formula_cell)["functions"] == ["SUM"]
    assert _meta(status_cell)["validation"]["values"] == ["Backlog", "In Progress", "Done"]
    assert _meta(named_cell)["named_ranges"] == ["TaskNames"]
    assert _meta(merged_cell)["merged_range"] == "F1:G1"


def test_builder_creates_csv_cartridge_with_searchable_cells(tmp_path):
    source = tmp_path / "people.csv"
    source.write_text("Name,Role\nAlice,Engineer\nBob,Designer\n", encoding="utf-8")
    output = tmp_path / "people.lun"

    asyncio.run(CartridgeBuilder(extract=False, embed=False).build(source, output))

    conn = sqlite3.connect(output)
    try:
        meta = dict(conn.execute("SELECT key, value FROM meta").fetchall())
        node_counts = dict(
            conn.execute("SELECT type, COUNT(*) FROM doc_nodes GROUP BY type").fetchall()
        )
        fts_hit = conn.execute(
            "SELECT rowid FROM nodes_fts WHERE nodes_fts MATCH 'Alice'"
        ).fetchone()

        assert meta["source_format"] == "csv"
        assert node_counts["table"] == 1
        assert node_counts["row"] == 3
        assert node_counts["cell"] == 6
        assert fts_hit is not None
    finally:
        conn.close()


def test_builder_creates_xlsx_cartridge_when_openpyxl_available(tmp_path):
    pytest.importorskip("openpyxl")
    source = tmp_path / "tasks.xlsx"
    _write_workbook(source)
    output = tmp_path / "tasks.lun"

    asyncio.run(CartridgeBuilder(extract=False, embed=False).build(source, output))

    conn = sqlite3.connect(output)
    try:
        meta = dict(conn.execute("SELECT key, value FROM meta").fetchall())
        formula_meta_json = conn.execute(
            "SELECT meta_json FROM doc_nodes WHERE type = 'cell' AND content LIKE '=SUM%'"
        ).fetchone()[0]
        formula_meta = json.loads(formula_meta_json)

        assert meta["source_format"] == "xlsx"
        assert formula_meta["formula"] == "=SUM(D2:D3)"
        assert formula_meta["formula_refs"] == ["Tasks!D2:D3"]
    finally:
        conn.close()


def test_builder_still_creates_markdown_cartridge(tmp_path):
    source = tmp_path / "note.md"
    source.write_text("# Note\n\nThis is still supported.", encoding="utf-8")
    output = tmp_path / "note.lun"

    asyncio.run(CartridgeBuilder(extract=False, embed=False).build(source, output))

    conn = sqlite3.connect(output)
    try:
        meta = dict(conn.execute("SELECT key, value FROM meta").fetchall())
        assert meta["source_format"] == "markdown"
        assert conn.execute("SELECT COUNT(*) FROM doc_nodes").fetchone()[0] > 0
    finally:
        conn.close()
