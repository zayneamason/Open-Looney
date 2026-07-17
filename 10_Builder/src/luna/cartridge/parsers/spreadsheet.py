"""XLSX parser for .lun cartridge builds."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .base import BaseParser
from .tabular import (
    column_letter,
    formula_functions,
    formula_refs,
    header_names,
    make_node,
    row_summary,
    stringify_value,
    value_type,
)


class SpreadsheetParser(BaseParser):
    """Parse XLSX workbooks into document/section/table/row/cell nodes."""

    def parse(self, source_path: Path) -> list[dict]:
        source_path = Path(source_path)
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX cartridge builds. "
                "Install with: pip install 'luna-engine[cartridge]'"
            ) from exc

        formula_wb = load_workbook(source_path, data_only=False, read_only=False)
        values_wb = load_workbook(source_path, data_only=True, read_only=False)

        nodes: list[dict] = [
            make_node(
                "document",
                None,
                None,
                0,
                {
                    "title": source_path.stem,
                    "source_format": "xlsx",
                    "sheet_count": len(formula_wb.worksheets),
                },
            )
        ]
        root_idx = 0

        named_ranges = self._named_ranges_by_sheet(formula_wb)
        for sheet_pos, ws in enumerate(formula_wb.worksheets):
            value_ws = values_wb[ws.title]
            bounds = self._used_bounds(ws, named_ranges.get(ws.title, {}))
            section_idx = len(nodes)
            nodes.append(
                make_node(
                    "section",
                    ws.title,
                    root_idx,
                    sheet_pos,
                    {
                        "sheet": ws.title,
                        "source_format": "xlsx",
                        "state": ws.sheet_state,
                    },
                )
            )
            if bounds is None:
                continue

            min_row, max_row, min_col, max_col = bounds
            header_row = self._first_nonempty_row(value_ws, min_row, max_row, min_col, max_col)
            if header_row is None:
                continue

            header_values = [
                value_ws.cell(header_row, col).value
                for col in range(min_col, max_col + 1)
            ]
            headers, headers_from_source = header_names(header_values)

            validations = self._validations_by_cell(ws, min_row, max_row, min_col, max_col)
            merged_ranges = self._merged_ranges_by_cell(ws, min_row, max_row, min_col, max_col)
            sheet_named_ranges = named_ranges.get(ws.title, {})

            table_idx = len(nodes)
            nodes.append(
                make_node(
                    "table",
                    None,
                    section_idx,
                    0,
                    {
                        "sheet": ws.title,
                        "range": f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}",
                        "row_count": max_row - header_row + 1,
                        "column_count": max_col - min_col + 1,
                        "headers_from_source": headers_from_source,
                        "headers": headers,
                    },
                )
            )

            row_pos = 0
            for row_num in range(header_row, max_row + 1):
                row_values = [
                    self._cell_display_value(ws.cell(row_num, col), value_ws.cell(row_num, col))
                    for col in range(min_col, max_col + 1)
                ]
                is_header = row_num == header_row
                row_meta = {
                    "sheet": ws.title,
                    "source_row": row_num,
                    "header": is_header,
                    "row_summary": "" if is_header else row_summary(headers, row_values),
                }
                row_idx = len(nodes)
                nodes.append(make_node("row", None, table_idx, row_pos, row_meta))

                for col in range(min_col, max_col + 1):
                    formula_cell = ws.cell(row_num, col)
                    value_cell = value_ws.cell(row_num, col)
                    col_pos = col - min_col
                    content = self._cell_display_value(formula_cell, value_cell)
                    meta = self._cell_meta(
                        ws.title,
                        formula_cell,
                        value_cell,
                        headers[col_pos],
                        validations.get(formula_cell.coordinate),
                        sheet_named_ranges.get(formula_cell.coordinate, []),
                        merged_ranges.get(formula_cell.coordinate),
                    )
                    nodes.append(make_node("cell", content, row_idx, col_pos, meta))

                row_pos += 1

        return nodes

    def _cell_display_value(self, formula_cell: Any, value_cell: Any) -> str:
        if self._is_formula(formula_cell.value):
            return stringify_value(value_cell.value if value_cell.value is not None else formula_cell.value)
        return stringify_value(formula_cell.value)

    def _cell_meta(
        self,
        sheet: str,
        formula_cell: Any,
        value_cell: Any,
        column_name: str,
        validation: dict | None,
        named_ranges: list[str],
        merged_range: str | None,
    ) -> dict:
        formula = formula_cell.value if self._is_formula(formula_cell.value) else None
        display_value = self._cell_display_value(formula_cell, value_cell)
        meta = {
            "sheet": sheet,
            "address": f"{sheet}!{formula_cell.coordinate}",
            "row": formula_cell.row,
            "column": formula_cell.column,
            "column_letter": column_letter(formula_cell.column),
            "column_name": column_name,
            "value": value_cell.value if formula else formula_cell.value,
            "display_value": display_value,
            "value_type": "formula" if formula else value_type(formula_cell.value),
            "number_format": getattr(formula_cell, "number_format", None),
        }
        if formula:
            meta["formula"] = formula
            meta["formula_refs"] = formula_refs(formula, sheet)
            meta["functions"] = formula_functions(formula)
        if validation:
            meta["validation"] = validation
        if named_ranges:
            meta["named_ranges"] = sorted(named_ranges)
        if merged_range:
            meta["merged_range"] = merged_range
        return meta

    def _used_bounds(self, ws: Any, named_ranges: dict[str, list[str]]) -> tuple[int, int, int, int] | None:
        rows: list[int] = []
        cols: list[int] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    rows.append(cell.row)
                    cols.append(cell.column)
        for merged in ws.merged_cells.ranges:
            rows.extend([merged.min_row, merged.max_row])
            cols.extend([merged.min_col, merged.max_col])
        for coord in named_ranges:
            cell = ws[coord]
            rows.append(cell.row)
            cols.append(cell.column)
        if not rows or not cols:
            return None
        return min(rows), max(rows), min(cols), max(cols)

    def _first_nonempty_row(
        self,
        ws: Any,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> int | None:
        for row_num in range(min_row, max_row + 1):
            values = [ws.cell(row_num, col).value for col in range(min_col, max_col + 1)]
            if any(value not in (None, "") for value in values):
                return row_num
        return None

    def _validations_by_cell(
        self,
        ws: Any,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> dict[str, dict]:
        validations: dict[str, dict] = {}
        for validation in ws.data_validations.dataValidation:
            serialized = self._serialize_validation(validation, ws.title)
            for row_num in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = ws.cell(row_num, col)
                    if cell.coordinate in validation.cells:
                        validations[cell.coordinate] = serialized
        return validations

    def _serialize_validation(self, validation: Any, sheet: str) -> dict:
        result = {
            "type": validation.type,
            "operator": validation.operator,
            "allow_blank": bool(validation.allow_blank),
            "formula1": validation.formula1,
            "formula2": validation.formula2,
            "prompt": validation.prompt,
            "error": validation.error,
        }
        if validation.type == "list" and validation.formula1:
            formula = validation.formula1
            if formula.startswith('"') and formula.endswith('"'):
                result["values"] = [item.strip() for item in formula.strip('"').split(",")]
            else:
                result["source_refs"] = formula_refs(f"={formula.lstrip('=')}", sheet)
        return result

    def _merged_ranges_by_cell(
        self,
        ws: Any,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> dict[str, str]:
        merged: dict[str, str] = {}
        for merged_range in ws.merged_cells.ranges:
            for row_num in range(max(min_row, merged_range.min_row), min(max_row, merged_range.max_row) + 1):
                for col in range(max(min_col, merged_range.min_col), min(max_col, merged_range.max_col) + 1):
                    coord = ws.cell(row_num, col).coordinate
                    merged[coord] = str(merged_range)
        return merged

    def _named_ranges_by_sheet(self, workbook: Any) -> dict[str, dict[str, list[str]]]:
        from openpyxl.utils.cell import range_boundaries

        by_sheet: dict[str, dict[str, list[str]]] = {}
        for defined_name in workbook.defined_names.values():
            try:
                destinations = list(defined_name.destinations)
            except Exception:
                continue
            for sheet_name, coord in destinations:
                if sheet_name not in workbook.sheetnames:
                    continue
                ws = workbook[sheet_name]
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(coord)
                except ValueError:
                    continue
                sheet_ranges = by_sheet.setdefault(sheet_name, {})
                for row_num in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_coord = ws.cell(row_num, col).coordinate
                        sheet_ranges.setdefault(cell_coord, []).append(defined_name.name)
        return by_sheet

    def _is_formula(self, value: Any) -> bool:
        return isinstance(value, str) and value.startswith("=")
